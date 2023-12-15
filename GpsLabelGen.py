from openpyxl.workbook import Workbook
from openpyxl import load_workbook
outputfile1 = open('GpsLabels.txt','w')

#Create a Workbook object
wb = load_workbook(filename = 'MachineNames.xlsx', data_only=True)

def generateLabels():
    #Opens Variables Worksheet
    ws = wb['all']

    #Grab Columns
    NucName                 = ws['a']
    mon1_flag               = ws['b'] 
    mon2_flag               = ws['c'] 
    mon3_flag               = ws['d']
    mon4_flag               = ws['e']    
    touch_flag              = ws['f']
    panelID                 = ws['g']
    orient                  = ws['h']    

    #I am creating an empty list for each column. then appending to list.
    #Probably a smarter way to do this but whatever.

    NucName_list            = []
    mon1_flag_list          = []
    mon2_flag_list          = []
    mon3_flag_list          = []
    mon4_flag_list          = []    
    touch_flag_list         = []
    panelID_list            = []
    orient_list             = []    

    for row in NucName[1:]:
        NucName_list.append(row.value)    
        
    for row in mon1_flag[1:]:
        mon1_flag_list.append(row.value)   

    for row in mon2_flag[1:]:
        mon2_flag_list.append(row.value)   

    for row in mon3_flag[1:]:
        mon3_flag_list.append(row.value)
        
    for row in mon4_flag[1:]:
        mon4_flag_list.append(row.value)       
        
    for row in touch_flag[1:]:
        touch_flag_list.append(row.value)
        
    for row in panelID[1:]:
        panelID_list.append(row.value)

    for row in orient[1:]:
        orient_list.append(row.value)            
        


    for i, item in enumerate(NucName_list):
        '''Horizontal Orientation'''
        if orient_list[i] == "H":
            '''Monitor 1'''
            if mon1_flag_list[i] is True:
                outputfile1.write(f'VIDEO 1\n')
                outputfile1.write(f'POWER 1\n')           
                '''Check if Touchscreen'''
                if touch_flag_list[i] is True:
                    outputfile1.write(f'TOUCH 1\n')
                      
            '''Monitor 2'''
            if mon2_flag_list[i] is True:
                outputfile1.write(f'VIDEO 2\n')
                outputfile1.write(f'POWER 2\n')           
                '''Check if Touchscreen'''
                if touch_flag_list[i] is True:
                    outputfile1.write(f'TOUCH2\n')

            '''Monitor 3'''
            if mon3_flag_list[i] is True:
                outputfile1.write(f'VIDEO 3\n')
                outputfile1.write(f'POWER 3\n')           
                '''Check if Touchscreen'''
                if touch_flag_list[i] is True:
                    outputfile1.write(f'TOUCH 3\n')

            '''Monitor 4'''
            if mon4_flag_list[i] is True:
                outputfile1.write(f'VIDEO 4\n')
                outputfile1.write(f'POWER 4\n')           
                '''Check if Touchscreen'''
                if touch_flag_list[i] is True:
                    outputfile1.write(f'{NucName_list[i]} TOUCH 4\n')

        elif orient_list[i] == "V":
            '''Vertical Orientation'''
            if mon1_flag_list[i] is True:
                outputfile1.write(f'VIDEO TOP\n')
                outputfile1.write(f'POWER TOP\n')           
                '''Check if Touchscreen'''
                if touch_flag_list[i] is True:
                    outputfile1.write(f'TOUCH TOP\n')
                      
            '''Monitor Middle'''
            if mon2_flag_list[i] is True:
                outputfile1.write(f'VIDEO MID\n')
                outputfile1.write(f'POWER MID\n')           
                '''Check if Touchscreen'''
                if touch_flag_list[i] is True:
                    outputfile1.write(f'TOUCH MID\n')

            '''Monitor Bottom'''
            if mon3_flag_list[i] is True:
                outputfile1.write(f'VIDEO BOT\n')
                outputfile1.write(f'POWER BOT\n')           
                '''Check if Touchscreen'''
                if touch_flag_list[i] is True:
                    outputfile1.write(f'TOUCH BOT\n')

        elif orient_list[i] == "S":
            '''Single Monitor'''
            if mon1_flag_list[i] is True:
                outputfile1.write(f'OVA VIDEO\n')
                outputfile1.write(f'OVA POWER\n')           
                '''Check if Touchscreen'''
                if touch_flag_list[i] is True:
                    outputfile1.write(f'OVA TOUCH\n')

        outputfile1.write(f'{NucName_list[i]} ETH\n')
        outputfile1.write(f'{NucName_list[i]} POWER\n')




     
generateLabels()